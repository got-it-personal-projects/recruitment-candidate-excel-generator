import base64
from io import BytesIO
from typing import Dict

from flask import render_template
from openpyxl import Workbook
from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import WHITE
from sendgrid.helpers.mail import Attachment, FileContent, FileName, FileType, Mail

from main import config, mail


class CandidateSpreadsheet:
    def __init__(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.column_dimensions["A"].width = 30
        self.sheet.column_dimensions["B"].width = 50
        self.row_count = 1

    def add_section(self, title: str, styles: Dict[str, Serialisable]):
        self.sheet.merge_cells(f"A{self.row_count}:B{self.row_count}")
        for k, v in styles.items():
            setattr(self.sheet[f"A{self.row_count}"], k, v)
        self.sheet[f"A{self.row_count}"] = title
        self.row_count += 1

    def add_details(self, data: dict):
        for k, v in data.items():
            self.sheet[f"A{self.row_count}"].alignment = Alignment(wrap_text=True)
            self.sheet[f"A{self.row_count}"] = k
            self.sheet[f"B{self.row_count}"].alignment = Alignment(wrap_text=True)
            self.sheet[f"B{self.row_count}"] = v
            self.row_count += 1

    def save_in_memory(self) -> BytesIO:
        output = BytesIO()
        self.workbook.save(output)
        return output


def save_candidate(candidate_data: dict) -> None:
    candidate_spreadsheet = CandidateSpreadsheet()
    main_section_header_styles = {
        "alignment": Alignment(horizontal="center"),
        "font": Font(bold=True, color=WHITE, size="18"),
        "fill": PatternFill(start_color="52BA71", fill_type="solid"),
    }
    sub_section_header_styles = {
        "alignment": Alignment(horizontal="center"),
        "font": Font(italic=True),
        "fill": PatternFill(start_color="EDEDED", fill_type="solid"),
    }

    personal_info_title = "Personal Information"
    personal_info_details = candidate_data[personal_info_title]
    candidate_spreadsheet.add_section(
        title=personal_info_title,
        styles=main_section_header_styles,
    )
    candidate_spreadsheet.add_details(personal_info_details)

    education_title = "Education"
    education_details = candidate_data[education_title]
    candidate_spreadsheet.add_section(
        title=education_title,
        styles=main_section_header_styles,
    )
    candidate_spreadsheet.add_details(education_details)

    cs_foundation_title = "Computer Science Foundation"
    cs_foundation_details = candidate_data[cs_foundation_title]
    candidate_spreadsheet.add_section(
        title=cs_foundation_title,
        styles=main_section_header_styles,
    )
    candidate_spreadsheet.add_details(cs_foundation_details)

    coding_sample_title = "Coding Samples"
    candidate_spreadsheet.add_section(
        title=coding_sample_title,
        styles=main_section_header_styles,
    )

    coding_sample_1_title = "PROJECT 1"
    coding_sample_1_details = candidate_data[coding_sample_title][coding_sample_1_title]
    candidate_spreadsheet.add_section(
        title=coding_sample_1_title,
        styles=sub_section_header_styles,
    )
    candidate_spreadsheet.add_details(coding_sample_1_details)

    coding_sample_2_title = "PROJECT 2"
    coding_sample_2_details = candidate_data[coding_sample_title][coding_sample_2_title]
    candidate_spreadsheet.add_section(
        title=coding_sample_2_title,
        styles=sub_section_header_styles,
    )
    candidate_spreadsheet.add_details(coding_sample_2_details)

    job_application_at_gotit_title = "Job Application at Got It"
    job_application_at_gotit_details = candidate_data[job_application_at_gotit_title]
    candidate_spreadsheet.add_section(
        title=job_application_at_gotit_title,
        styles=main_section_header_styles,
    )
    candidate_spreadsheet.add_details(job_application_at_gotit_details)

    other_info_title = "Other Information"
    candidate_spreadsheet.add_section(
        title=other_info_title,
        styles=main_section_header_styles,
    )

    english_title = "ENGLISH"
    english_details = candidate_data[other_info_title][english_title]
    candidate_spreadsheet.add_section(
        title=english_title,
        styles=sub_section_header_styles,
    )
    candidate_spreadsheet.add_details(english_details)

    future_plans_title = "FUTURE PLANS"
    future_plans_details = candidate_data[other_info_title][future_plans_title]
    candidate_spreadsheet.add_section(
        title=future_plans_title,
        styles=sub_section_header_styles,
    )
    candidate_spreadsheet.add_details(future_plans_details)

    references_title = "REFERENCES"
    references_details = candidate_data[other_info_title][references_title]
    candidate_spreadsheet.add_section(
        title=references_title,
        styles=sub_section_header_styles,
    )
    candidate_spreadsheet.add_details(references_details)

    output = candidate_spreadsheet.save_in_memory()

    _send_mail_with_candidate_excel_to_recruiter(
        subject=f"{personal_info_details['Full Name']}'s Application Form",
        content=render_template("email/recruitment_notification.txt"),
        excel_filename=f"{personal_info_details['Full Name']}.xlsx",
        excel_data=output.getvalue(),
    )


def _send_mail_with_candidate_excel_to_recruiter(
    *,
    subject: str,
    content: str,
    excel_filename: str,
    excel_data: bytes,
):
    message = Mail(
        from_email=config.SENDER_EMAIL,
        to_emails=config.RECIPIENT_EMAILS,
        subject=subject,
        plain_text_content=content,
    )
    encode_excel_data = base64.b64encode(excel_data).decode()
    attachment = Attachment(
        file_content=FileContent(encode_excel_data),
        file_name=FileName(excel_filename),
        file_type=FileType("application/vnd.ms-excel"),
    )
    message.attachment = attachment

    mail.send(message)
